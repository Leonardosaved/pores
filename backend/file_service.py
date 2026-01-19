import os
import cv2
import numpy as np
import openpyxl
import json
from openpyxl import Workbook
from typing import List, Dict, Set

def get_analyzed_images(folder_path: str) -> Set[str]:
    """
    Reads the Excel file and returns a set of image names that have at least one ROI entry.
    """
    filepath = os.path.join(folder_path, "roi_measurements.xlsx")
    analyzed_files = set()

    if not os.path.exists(filepath):
        return analyzed_files

    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)
        sheet = workbook.active
        # Assuming 'image_name' is the first column
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                analyzed_files.add(row[0])
    except Exception:
        # If the file is corrupt or unreadable, return an empty set
        return set()

    return analyzed_files

def update_roi_notes(folder_path: str, image_name: str, selection_number: int, notes: str):
    """
    Updates the notes for an existing ROI in-place without creating a new version row.
    This is for notes-only updates that don't change geometry.
    """
    filepath = os.path.join(folder_path, "roi_measurements.xlsx")
    
    if not os.path.exists(filepath):
        return  # No Excel file to update
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Get header
        headers = [cell.value for cell in sheet[1]]
        header_keys = [str(name).strip().lower() if name else "" for name in headers]
        
        # Find notes column
        notes_col = None
        for idx, key in enumerate(header_keys, start=1):
            if key == "notes":
                notes_col = idx
                break
        
        if notes_col is None:
            # Add notes column if it doesn't exist
            notes_col = len(headers) + 1
            sheet.cell(row=1, column=notes_col, value="notes")
        
        # Find the latest version of this ROI for this image
        target_row = None
        target_version = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if (row[0].value == image_name and 
                row[1].value == selection_number):
                # Keep track of the highest version
                version = row[2].value or 1
                if version > target_version:
                    target_version = version
                    target_row = row_idx
        
        # Update notes in the latest version row
        if target_row:
            sheet.cell(row=target_row, column=notes_col, value=notes)
            workbook.save(filepath)
    
    except Exception as e:
        print(f"Warning: Could not update ROI notes: {e}")

def save_roi_to_excel(folder_path: str, data: Dict):
    """
    Appends a new row with ROI data to the Excel file.
    Creates the file and header if it doesn't exist.
    """
    filepath = os.path.join(folder_path, "roi_measurements.xlsx")

    header = [
        "image_name", "selection_number", "version", "scale_px_per_um",
        "scale_um", "scale_bar_x1", "scale_bar_y1", "scale_bar_x2", "scale_bar_y2",
        "area_um2", "area_px2", "points_json", "notes", "overlay_file"
    ]

    try:
        if not os.path.exists(filepath):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(header)
            workbook.save(filepath)

        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Ensure header has notes/overlay columns if file was created with an older schema.
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if header_row:
            headers = [cell.value for cell in header_row]
            if "notes" not in headers:
                sheet.cell(row=1, column=len(headers) + 1, value="notes")
                headers.append("notes")
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=len(headers), value="")
            if "overlay_file" not in headers:
                sheet.cell(row=1, column=len(headers) + 1, value="overlay_file")
                headers.append("overlay_file")
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=len(headers), value="")

        headers = [cell.value for cell in sheet[1]]
        header_keys = [str(name).strip().lower() if name else "" for name in headers]
        data_map = {
            "image_name": data["image_name"],
            "selection_number": data["selection_number"],
            "version": data.get("version", 1),
            "scale_px_per_um": data["scale_px_per_um"],
            "scale_um": data.get("scale_um", 0),
            "scale_bar_x1": data.get("scale_bar", {}).get("x1") if data.get("scale_bar") else None,
            "scale_bar_y1": data.get("scale_bar", {}).get("y1") if data.get("scale_bar") else None,
            "scale_bar_x2": data.get("scale_bar", {}).get("x2") if data.get("scale_bar") else None,
            "scale_bar_y2": data.get("scale_bar", {}).get("y2") if data.get("scale_bar") else None,
            "area_um2": data["area_um2"],
            "area_px2": data["area_px2"],
            "points_json": json.dumps(data.get("points", [])),
            "notes": data.get("notes", ""),
            "overlay_file": data["overlay_file"],
        }
        # If modifying existing ROI, append a new row with incremented version (don't delete old rows)
        # This creates a history of ROI modifications
        row_to_add = [data_map.get(key, None) for key in header_keys]
        sheet.append(row_to_add)
        workbook.save(filepath)

    except PermissionError:
        raise PermissionError("Could not write to Excel. Please close the file and try again.")
    except Exception as e:
        raise IOError(f"Failed to write to Excel file: {e}")

def save_scale_bar(folder_path: str, image_name: str, scale_bar: Dict, scale_um: float):
    """
    Saves scale bar data independently to a JSON config file.
    This persists scale bar settings even without ROIs.
    """
    import json
    config_file = os.path.join(folder_path, ".pore_analyzer_config.json")
    
    try:
        # Validate scale_um before saving
        if scale_um and (scale_um < 1 or scale_um > 10000):
            scale_um = 100  # Default to 100 if invalid
        
        # Load existing config
        config = {}
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                config = json.load(f)
        
        # Update scale bar for this image
        if "scale_bars" not in config:
            config["scale_bars"] = {}
        
        config["scale_bars"][image_name] = {
            "scale_bar": scale_bar,
            "scale_um": scale_um
        }
        
        # Save back
        with open(config_file, 'w') as f:
            json.dump(config, f, indent=2)
    except Exception as e:
        print(f"Warning: Could not save scale bar config: {e}")

def load_scale_bar(folder_path: str, image_name: str) -> Dict:
    """
    Loads scale bar data from config file for a specific image.
    """
    import json
    config_file = os.path.join(folder_path, ".pore_analyzer_config.json")
    
    if not os.path.exists(config_file):
        return {"scaleBar": None, "scaleUm": 0}
    
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        
        if "scale_bars" in config and image_name in config["scale_bars"]:
            data = config["scale_bars"][image_name]
            scale_um = data.get("scale_um", 0)
            # Validate scale_um is in reasonable range
            if scale_um and (scale_um < 1 or scale_um > 10000):
                scale_um = 100  # Default to 100 if corrupted
            return {
                "scaleBar": data.get("scale_bar"),
                "scaleUm": scale_um
            }
    except Exception as e:
        print(f"Warning: Could not load scale bar config: {e}")
    
    return {"scaleBar": None, "scaleUm": 0}

def load_roi_data(folder_path: str, image_name: str) -> Dict:
    """
    Loads all ROI data for a specific image from the Excel file.
    Returns the latest version of each ROI.
    Also loads scale bar data from config if available.
    """
    filepath = os.path.join(folder_path, "roi_measurements.xlsx")
    roi_data = {"rois": [], "scaleBar": None, "scaleUm": 0}

    if not os.path.exists(filepath):
        # No Excel file yet, try to load scale bar from config
        config_data = load_scale_bar(folder_path, image_name)
        return {**roi_data, **config_data}

    try:
        import json
        workbook = openpyxl.load_workbook(filepath, read_only=True)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        header_map = {str(name).strip().lower(): idx for idx, name in enumerate(header_row) if name}

        def get_value(row, key, fallback_idx=None):
            idx = header_map.get(key)
            if idx is None:
                idx = fallback_idx
            if idx is None or idx >= len(row):
                return None
            return row[idx]
        
        roi_dict = {}  # Key: selection_number, Value: latest version row
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if get_value(row, "image_name", 0) == image_name:
                selection_number = get_value(row, "selection_number", 1)
                version = get_value(row, "version", 2) or 1
                
                # Keep only the latest version
                if selection_number not in roi_dict or roi_dict[selection_number]["version"] < version:
                    roi_dict[selection_number] = {
                        "version": version,
                        "scale_px_per_um": get_value(row, "scale_px_per_um", 3),
                        "scale_um": get_value(row, "scale_um", 4),
                        "scale_bar_x1": get_value(row, "scale_bar_x1", 5),
                        "scale_bar_y1": get_value(row, "scale_bar_y1", 6),
                        "scale_bar_x2": get_value(row, "scale_bar_x2", 7),
                        "scale_bar_y2": get_value(row, "scale_bar_y2", 8),
                        "area_um2": get_value(row, "area_um2", 9),
                        "area_px2": get_value(row, "area_px2", 10),
                        "points_json": get_value(row, "points_json", 11),
                        "notes": (get_value(row, "notes", None) if "notes" in header_map else "") or "",
                    }
        
        # Convert dict to list and parse points
        for roi_id, roi_data_row in roi_dict.items():
            try:
                points = json.loads(roi_data_row["points_json"]) if isinstance(roi_data_row["points_json"], str) else []
            except:
                points = []
            
            roi_data["rois"].append({
                "id": roi_id,
                "version": roi_data_row["version"],
                "points": points,
                "areaPx2": roi_data_row["area_px2"],
                "areaUm2": roi_data_row["area_um2"],
                "notes": roi_data_row.get("notes", ""),
            })
            
            # Set scale bar from first ROI (same for all ROIs in same image)
            if roi_data["scaleBar"] is None and roi_data_row["scale_bar_x1"] is not None:
                roi_data["scaleBar"] = {
                    "x1": roi_data_row["scale_bar_x1"],
                    "y1": roi_data_row["scale_bar_y1"],
                    "x2": roi_data_row["scale_bar_x2"],
                    "y2": roi_data_row["scale_bar_y2"],
                }
                # Validate scale_um from Excel
                scale_um_val = roi_data_row["scale_um"] or 0
                if scale_um_val and (scale_um_val < 1 or scale_um_val > 10000):
                    scale_um_val = 100  # Default to 100 if corrupted
                roi_data["scaleUm"] = scale_um_val
        
        # If no scale bar found in ROIs, try config file
        if roi_data["scaleBar"] is None:
            config_data = load_scale_bar(folder_path, image_name)
            roi_data["scaleBar"] = config_data.get("scaleBar")
            roi_data["scaleUm"] = config_data.get("scaleUm", 0)

    except Exception as e:
        print(f"Error loading ROI data: {e}")
        # Still try to get scale bar from config even if Excel fails
        config_data = load_scale_bar(folder_path, image_name)
        return {**{"rois": [], "scaleBar": None, "scaleUm": 0}, **config_data}

    return roi_data

def save_overlay_image(folder_path: str, image_name: str, selection_number: int, points: List[Dict], version: int = 1):
    """
    Draws the ROI polygon on the original image and saves it as a PNG.
    Filename format: image_roinumber_vversion.png
    """
    output_dir = os.path.join(folder_path, "_roi_overlays")
    os.makedirs(output_dir, exist_ok=True)

    original_image_path = os.path.join(folder_path, image_name)
    base_name, _ = os.path.splitext(image_name)
    output_filename = f"{base_name}_{selection_number}_v{version}.png"
    output_path = os.path.join(output_dir, output_filename)

    try:
        img = cv2.imread(original_image_path)
        if img is None:
            raise IOError("Could not read original image.")

        # Convert points to the format cv2.polylines needs
        pts = np.array([[p['x'], p['y']] for p in points], np.int32)
        pts = pts.reshape((-1, 1, 2))

        # Draw the polygon
        cv2.polylines(img, [pts], isClosed=True, color=(0, 0, 255), thickness=2)

        # Add a label with version
        label = f"ROI {selection_number} v{version}"
        label_pos = (pts[0][0][0], pts[0][0][1] - 10)
        cv2.putText(img, label, label_pos, cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

        cv2.imwrite(output_path, img)
        return output_filename

    except Exception as e:
        raise IOError(f"Failed to save overlay image: {e}")

def save_notes(folder_path: str, image_name: str, notes: str):
    """
    Saves analysis notes for an image to a JSON config file.
    Notes persist independently of ROI data.
    """
    config_file = os.path.join(folder_path, ".pore_analyzer_notes.json")
    
    try:
        # Load existing notes
        notes_data = {}
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                notes_data = json.load(f)
        
        # Update or add notes for this image
        notes_data[image_name] = notes
        
        # Save back
        with open(config_file, 'w') as f:
            json.dump(notes_data, f, indent=2)
    except Exception as e:
        print(f"Warning: Could not save notes: {e}")

def load_notes(folder_path: str, image_name: str) -> str:
    """
    Loads analysis notes for a specific image.
    """
    config_file = os.path.join(folder_path, ".pore_analyzer_notes.json")
    
    if not os.path.exists(config_file):
        return ""
    
    try:
        with open(config_file, 'r') as f:
            notes_data = json.load(f)
        
        return notes_data.get(image_name, "")
    except Exception as e:
        print(f"Warning: Could not load notes: {e}")
        return ""

def delete_image_analysis(folder_path: str, image_name: str):
    """
    Deletes all ROI data for a specific image:
    - Removes all rows from the Excel file for that image
    - Deletes all overlay images for that image
    - Removes scale bar config for that image
    """
    # Delete from Excel
    excel_path = os.path.join(folder_path, "roi_measurements.xlsx")
    if os.path.exists(excel_path):
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
            
            # Find and delete all rows for this image (iterate backwards to avoid index issues)
            rows_to_delete = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if row[0].value == image_name:
                    rows_to_delete.append(row_idx)
            
            # Delete rows in reverse order
            for row_idx in reversed(rows_to_delete):
                sheet.delete_rows(row_idx, 1)
            
            workbook.save(excel_path)
        except Exception as e:
            raise IOError(f"Failed to delete Excel data: {e}")
    
    # Delete overlay images
    overlay_dir = os.path.join(folder_path, "_roi_overlays")
    if os.path.exists(overlay_dir):
        try:
            base_name = os.path.splitext(image_name)[0]
            for filename in os.listdir(overlay_dir):
                if filename.startswith(base_name + "_") and filename.endswith(".png"):
                    file_path = os.path.join(overlay_dir, filename)
                    os.remove(file_path)
        except Exception as e:
            raise IOError(f"Failed to delete overlay images: {e}")
    
    # Delete scale bar config for this image
    config_file = os.path.join(folder_path, ".pore_analyzer_config.json")
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
            
            # Remove scale bar entry for this image
            if "scale_bars" in config and image_name in config["scale_bars"]:
                del config["scale_bars"][image_name]
                
                # Save updated config
                with open(config_file, 'w') as f:
                    json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Warning: Could not delete scale bar config: {e}")
